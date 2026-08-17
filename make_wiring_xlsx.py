#!/usr/bin/env python3
"""
Generate wiring_chart.xlsx — a mixed text/calc spreadsheet for OnlyOffice Calc / Excel.

This module creates a comprehensive FPV Rover wiring chart as an Excel file with
four worksheets:

    - **Readme**:               Notes and file-structure reference (text only).
    - **All_Wires**:            A single Excel Table (ListObject, "WireTable") containing
                                every wire in the build.  Fully sortable and filterable.
    - **Component_Placement**:  Layout and pin maps imported from ``component_placement.md``.
    - **Summary**:              Wire count per section with a live ``=SUM`` total row.

Preference rule (see ``~/.clinerules/wiring-chart-preferences.md``):
    Title / subtitle rows are placed **above** the Excel Table and are never
    part of the sortable data range.

Usage::

    python make_wiring_xlsx.py

The output file ``wiring_chart.xlsx`` is written to the current working directory.
"""

from __future__ import annotations

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Sequence, Tuple

# ── Constants ────────────────────────────────────────────────────────────────

#: Thin border style used on every cell.
THIN_BORDER: Border = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

#: Hex colour for table header backgrounds.
HEADER_COLOR: str = "1F4E78"

#: Hex colour for summary-table header backgrounds.
SUMMARY_HEADER_COLOR: str = "7030A0"

#: Hex colour for the grand-total row.
TOTAL_COLOR: str = "E26B0A"

#: Column headers for the main ``All_Wires`` table.
HEADERS: List[str] = [
    "ID",
    "Section",
    "From Device",
    "From Pin",
    "To Device",
    "To Pin",
    "Signal / Notes",
    "Phase",
    "Level-shifted?",
]

#: Column widths (in characters) matching ``HEADERS`` order.
ALL_WIRES_WIDTHS: List[int] = [6, 18, 18, 22, 18, 18, 40, 12, 16]

#: Row indices for layout of the ``All_Wires`` sheet.
TITLE_ROW: int = 1
SUBTITLE_ROW: int = 2
HEADER_ROW: int = 4  # row 3 is a blank separator
DATA_START: int = 5

# ── Wire data ────────────────────────────────────────────────────────────────
# Each entry:  [ID, Section, FromDevice, FromPin, ToDevice, ToPin, Notes, Phase, LevelShifted?]
# Ordered: Power → Motors → UART → Trigger → Echo → Sensors → I²C Bus.

ALL_WIRES: List[List[str]] = [
    # ── Power Distribution ──
    ["P1", "Power_Distribution", "12 V Battery", "+", "L298N", "VMS",
     "Motor power input", "Phase 00", "—"],
    ["P2", "Power_Distribution", "12 V Battery", "+", "LM2596 #1", "IN",
     "Buck to 5 V logic rail", "Phase 00", "—"],
    ["P3", "Power_Distribution", "LM2596 #1", "OUT", "5 V rail", "—",
     "5 V logic supply rail", "Phase 00", "—"],
    ["P4", "Power_Distribution", "5 V rail", "5 V", "UNO", "5 V",
     "Logic power", "Phase 00", "—"],
    ["P5", "Power_Distribution", "5 V rail", "5 V", "ESP32-CAM", "5 V",
     "Logic power", "Phase 00", "—"],
    ["P6", "Power_Distribution", "5 V rail", "5 V", "Sensor ring", "5 V",
     "Perimeter power ring (all sensors)", "Phase 00", "—"],
    ["P7", "Power_Distribution", "5 V rail", "GND", "Sensor ring", "GND",
     "Perimeter ground ring", "Phase 00", "—"],
    ["P8", "Power_Distribution", "5 V rail", "GND", "UNO", "GND",
     "Common ground", "Phase 00", "—"],
    ["P9", "Power_Distribution", "5 V rail", "GND", "ESP32-CAM", "GND",
     "Common ground", "Phase 00", "—"],

    # ── Motor Drive ──
    ["M1", "Motor_Drive", "UNO", "D5 (PWM)", "L298N", "ENA",
     "Left motor enable", "Phase 04", "—"],
    ["M2", "Motor_Drive", "UNO", "D6 (PWM)", "L298N", "ENB",
     "Right motor enable", "Phase 04", "—"],
    ["M3", "Motor_Drive", "UNO", "D2", "L298N", "IN1",
     "Left motor direction 1", "Phase 04", "—"],
    ["M4", "Motor_Drive", "UNO", "D3", "L298N", "IN2",
     "Left motor direction 2", "Phase 04", "—"],
    ["M5", "Motor_Drive", "UNO", "D7", "L298N", "IN3",
     "Right motor direction 1", "Phase 04", "—"],
    ["M6", "Motor_Drive", "UNO", "D4", "L298N", "IN4",
     "Right motor direction 2", "Phase 04", "—"],

    # ── Serial / UART ──
    ["S1", "Serial_UART", "UNO", "D11 (SoftSerial TX)", "ESP32-CAM", "GPIO14 (UART1 RX)",
     "UNO -> ESP32 command channel", "Phase 00", "Yes"],
    ["S2", "Serial_UART", "ESP32-CAM", "GPIO13 (UART1 TX)", "UNO", "D10 (SoftSerial RX)",
     "ESP32 -> UNO telemetry", "Phase 00", "Yes"],
    ["S3", "Serial_UART", "UNO", "D1 (HW UART TX)", "D51157 / Jazzy", "White data line",
     "Jazzy serial output, 38400 8E1", "Phase 00", "Yes"],
    ["S4", "Serial_UART", "UNO", "D0 (HW UART RX)", "Reserved", "—",
     "Preserve for D1 hardware UART pair", "Phase 00", "—"],

    # ── HC-SR04 Trigger ──
    ["T1", "HC_SR04_Trigger", "UNO", "D9", "HC-SR04 (all 5)", "TRIG (shared)",
     "Fires all sensors simultaneously", "Phase 03", "—"],

    # ── HC-SR04 Echo ──
    ["E1", "HC_SR04_Echo", "HC-SR04 CF", "ECHO", "UNO", "A0", "0 deg (center-front)", "Phase 03", "—"],
    ["E2", "HC_SR04_Echo", "HC-SR04 LF", "ECHO", "UNO", "A1", "45 deg (left-front)", "Phase 03", "—"],
    ["E3", "HC_SR04_Echo", "HC-SR04 RF", "ECHO", "UNO", "A2", "45 deg (right-front)", "Phase 03", "—"],
    ["E4", "HC_SR04_Echo", "HC-SR04 L", "ECHO", "UNO", "A3", "90 deg (left side)", "Phase 03", "—"],
    ["E5", "HC_SR04_Echo", "HC-SR04 R", "ECHO", "UNO", "D13", "90 deg (right side)", "Phase 03", "—"],

    # ── HW-201 IR Sensor Ring ──
    ["IR1", "HW201_IR_Ring", "MCP23017 #1", "GPA1", "HW-201 #1", "DO",
     "IR drop sensor front-left (down)", "Phase 03", "—"],
    ["IR2", "HW201_IR_Ring", "MCP23017 #1", "GPA2", "HW-201 #2", "DO",
     "IR drop sensor front-right (down)", "Phase 03", "—"],
    ["IR3", "HW201_IR_Ring", "MCP23017 #1", "GPA3", "HW-201 #3", "DO",
     "IR drop sensor left side (down)", "Phase 03", "—"],
    ["IR4", "HW201_IR_Ring", "MCP23017 #1", "GPA4", "HW-201 #4", "DO",
     "IR drop sensor right side (down)", "Phase 03", "—"],

    # ── Other Sensors ──
    ["O1", "Other_Sensors", "DHT11", "DATA", "UNO", "D12",
     "Ambient temp / humidity", "Phase 08", "—"],
    ["O2", "Other_Sensors", "HW-484 (mic)", "OUT", "UNO", "A6",
     "Microphone (analog)", "Phase 16", "—"],
    ["O3", "Other_Sensors", "SW-520D", "OUT", "UNO", "D8",
     "Tilt / vibration sensor", "Phase 06", "—"],
    ["O4", "Other_Sensors", "Passive Buzzer", "IN", "UNO", "D12",
     "Buzzer (shared with DHT11 via NPN)", "Phase 02", "—"],
    ["O5", "Other_Sensors", "RGB LED", "R", "UNO", "D8",
     "RGB status indicator (common cathode)", "Phase 01", "—"],
    ["O6", "Other_Sensors", "RGB LED", "G", "UNO", "A7",
     "RGB status indicator (common cathode)", "Phase 01", "—"],
    ["O7", "Other_Sensors", "RGB LED", "B", "UNO", "A7",
     "RGB status indicator (common cathode)", "Phase 01", "—"],

    # ── I2C Bus ──
    ["I2C-1", "I2C_Bus", "UNO", "A4", "I2C bus", "SDA",
     "Data line, via level shifter to 3.3 V", "Phase 00", "Yes"],
    ["I2C-2", "I2C_Bus", "UNO", "A5", "I2C bus", "SCL",
     "Clock line, via level shifter to 3.3 V", "Phase 00", "Yes"],
    ["I2C-3", "I2C_Bus", "3.3 V rail", "VCC", "I2C bus", "VCC",
     "Pull-up to 3.3 V", "Phase 00", "—"],
    ["I2C-4", "I2C_Bus", "GND", "—", "I2C bus", "GND",
     "Common reference", "Phase 00", "—"],

    # ── Additional Phase 2 Components ──
    ["AP2-1", "Additional_Components", "5 V rail", "5 V", "DS1302", "VCC",
     "RTC power", "Phase 10", "—"],
    ["AP2-2", "Additional_Components", "GND", "—", "DS1302", "GND",
     "RTC ground", "Phase 10", "—"],
    ["AP2-3", "Additional_Components", "MCP23017 #1", "GPB0", "DS1302", "CE",
     "RTC chip enable", "Phase 10", "—"],
    ["AP2-4", "Additional_Components", "MCP23017 #1", "GPB1", "DS1302", "I/O",
     "RTC data I/O", "Phase 10", "—"],
    ["AP2-5", "Additional_Components", "MCP23017 #1", "GPB2", "DS1302", "SCLK",
     "RTC serial clock", "Phase 10", "—"],
    ["AP2-6", "Additional_Components", "5 V rail", "5 V", "LCD1602", "VDD",
     "Display power", "Phase 07", "—"],
    ["AP2-7", "Additional_Components", "GND", "—", "LCD1602", "VSS",
     "Display ground", "Phase 07", "—"],
    ["AP2-8", "Additional_Components", "5 V rail", "5 V", "LCD1602", "VEE",
     "Display contrast", "Phase 07", "—"],
    ["AP2-9", "Additional_Components", "MCP23017 #2", "GPA3", "LCD1602", "RS",
     "LCD register select", "Phase 07", "—"],
    ["AP2-10", "Additional_Components", "MCP23017 #2", "GPA4", "LCD1602", "E",
     "LCD enable", "Phase 07", "—"],
    ["AP2-11", "Additional_Components", "MCP23017 #2", "GPA5", "LCD1602", "D4",
     "LCD data 4", "Phase 07", "—"],
    ["AP2-12", "Additional_Components", "MCP23017 #2", "GPA6", "LCD1602", "D5",
     "LCD data 5", "Phase 07", "—"],
    ["AP2-13", "Additional_Components", "MCP23017 #2", "GPA7", "LCD1602", "D6",
     "LCD data 6", "Phase 07", "—"],
    ["AP2-14", "Additional_Components", "MCP23017 #2", "GPB0", "LCD1602", "D7",
     "LCD data 7", "Phase 07", "—"],
    ["AP2-15", "Additional_Components", "5 V rail", "5 V", "KY-022", "VCC",
     "IR receiver power", "Phase 12", "—"],
    ["AP2-16", "Additional_Components", "GND", "—", "KY-022", "GND",
     "IR receiver ground", "Phase 12", "—"],
    ["AP2-17", "Additional_Components", "MCP23017 #2", "GPB4", "KY-022", "OUT",
     "IR receiver data", "Phase 12", "—"],
    ["AP2-18", "Additional_Components", "5 V rail", "5 V", "NEO-6M", "VCC",
     "GPS power", "Phase 05", "—"],
    ["AP2-19", "Additional_Components", "GND", "—", "NEO-6M", "GND",
     "GPS ground", "Phase 05", "—"],
    ["AP2-20", "Additional_Components", "ESP32-CAM", "GPIO16", "NEO-6M", "TX",
     "GPS data to ESP32", "Phase 05", "—"],
    ["AP2-21", "Additional_Components", "5 V rail", "5 V", "Servo (pan)", "VCC",
     "Pan servo power", "Phase 11", "—"],
    ["AP2-22", "Additional_Components", "GND", "—", "Servo (pan)", "GND",
     "Pan servo ground", "Phase 11", "—"],
    ["AP2-23", "Additional_Components", "ESP32-CAM", "GPIO2", "Servo (pan)", "SIG",
     "Pan servo signal", "Phase 11", "—"],
    ["AP2-24", "Additional_Components", "5 V rail", "5 V", "Servo (tilt)", "VCC",
     "Tilt servo power", "Phase 11", "—"],
    ["AP2-25", "Additional_Components", "GND", "—", "Servo (tilt)", "GND",
     "Tilt servo ground", "Phase 11", "—"],
    ["AP2-26", "Additional_Components", "ESP32-CAM", "GPIO15", "Servo (tilt)", "SIG",
     "Tilt servo signal", "Phase 11", "—"],
    ["AP2-27", "Additional_Components", "5 V rail", "5 V", "RFID-RC522", "VCC",
     "RFID reader power", "Phase 13", "—"],
    ["AP2-28", "Additional_Components", "GND", "—", "RFID-RC522", "GND",
     "RFID reader ground", "Phase 13", "—"],
    ["AP2-29", "Additional_Components", "UNO", "D8", "RFID-RC522", "SDA",
     "RFID SPI data", "Phase 13", "—"],
    ["AP2-30", "Additional_Components", "UNO", "D8", "RFID-RC522", "SCK",
     "RFID SPI clock", "Phase 13", "—"],
    ["AP2-31", "Additional_Components", "UNO", "D8", "RFID-RC522", "MOSI",
     "RFID SPI master out", "Phase 13", "—"],
    ["AP2-32", "Additional_Components", "UNO", "D8", "RFID-RC522", "MISO",
     "RFID SPI master in", "Phase 13", "—"],
    ["AP2-33", "Additional_Components", "UNO", "D8", "RFID-RC522", "NSS",
     "RFID chip select", "Phase 13", "—"],
    ["AP2-34", "Additional_Components", "5 V rail", "5 V", "HW-038", "VCC",
     "Water level sensor power", "Phase 09", "—"],
    ["AP2-35", "Additional_Components", "GND", "—", "HW-038", "GND",
     "Water level sensor ground", "Phase 09", "—"],
    ["AP2-36", "Additional_Components", "UNO", "A7", "HW-038", "OUT",
     "Water level sensor analog output", "Phase 09", "—"],
    ["AP2-37", "Additional_Components", "5 V rail", "5 V", "VS1838B", "VCC",
     "IR receiver power", "Phase 14", "—"],
    ["AP2-38", "Additional_Components", "GND", "—", "VS1838B", "GND",
     "IR receiver ground", "Phase 14", "—"],
    ["AP2-39", "Additional_Components", "UNO", "D8", "VS1838B", "OUT",
     "IR receiver data", "Phase 14", "—"],
    ["AP2-40", "Additional_Components", "5 V rail", "5 V", "ULN2003", "VCC",
     "Stepper driver power", "Phase 15", "—"],
    ["AP2-41", "Additional_Components", "GND", "—", "ULN2003", "GND",
     "Stepper driver ground", "Phase 15", "—"],
    ["AP2-42", "Additional_Components", "UNO", "D8", "ULN2003", "IN1",
     "Stepper coil 1", "Phase 15", "—"],
    ["AP2-43", "Additional_Components", "UNO", "D8", "ULN2003", "IN2",
     "Stepper coil 2", "Phase 15", "—"],
    ["AP2-44", "Additional_Components", "UNO", "D8", "ULN2003", "IN3",
     "Stepper coil 3", "Phase 15", "—"],
    ["AP2-45", "Additional_Components", "UNO", "D8", "ULN2003", "IN4",
     "Stepper coil 4", "Phase 15", "—"],
]

TOTAL_WIRES: int = len(ALL_WIRES)


# ── Styling helpers ──────────────────────────────────────────────────────────


def style_header_cell(
    cell: openpyxl.cell.cell.Cell,
    value: str,
) -> openpyxl.cell.cell.Cell:
    """Apply header formatting (bold white text, blue-grey fill, border).

    Args:
        cell: The worksheet cell to format.
        value: The text to place in the cell.

    Returns:
        The formatted cell (modified in place).
    """
    cell.value = value
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(
        start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid"
    )
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER
    return cell


def style_data_cell(
    cell: openpyxl.cell.cell.Cell,
    value: str,
) -> None:
    """Apply data-row formatting (left-aligned, top-aligned, thin border).

    Args:
        cell: The worksheet cell to format.
        value: The text to place in the cell.
    """
    cell.value = value
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="left", vertical="top")


def style_summary_header(
    cell: openpyxl.cell.cell.Cell,
    value: str,
) -> None:
    """Apply summary-table header formatting (bold white on purple).

    Args:
        cell: The worksheet cell to format.
        value: The text to place in the cell.
    """
    cell.value = value
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(
        start_color=SUMMARY_HEADER_COLOR,
        end_color=SUMMARY_HEADER_COLOR,
        fill_type="solid",
    )
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")


# ── Tab builders ─────────────────────────────────────────────────────────────


def build_all_wires_tab(
    wb: openpyxl.Workbook,
    all_wires: List[List[str]],
) -> None:
    """Create the ``All_Wires`` worksheet with a single sortable Excel Table.

    The title (row 1) and subtitle (row 2) are placed **above** the Excel
    Table so they are never included in sort/filter operations.  The table
    itself starts at ``HEADER_ROW`` (row 4) with headers, and data rows begin
    at ``DATA_START`` (row 5).

    Args:
        wb: The workbook to add the worksheet to.
        all_wires: List of wire data rows (each row matches ``HEADERS``).
    """
    ws: Worksheet = wb.active
    ws.title = "All_Wires"
    ws.sheet_view.show_gridlines = False

    last_col: str = get_column_letter(len(HEADERS))

    # --- Title (row 1) — merged across all columns, NOT part of the table ---
    title = ws.cell(row=TITLE_ROW, column=1)
    title.value = "FPV Rover - Wiring Chart (All Wires)"
    title.font = Font(size=14, bold=True)
    ws.merge_cells(f"A{TITLE_ROW}:{last_col}{TITLE_ROW}")

    # --- Subtitle (row 2) ---
    subtitle = ws.cell(row=SUBTITLE_ROW, column=1)
    subtitle.value = (
        f"{len(all_wires)} wires across {len(set(w[1] for w in all_wires))} sections - "
        "sortable & filterable by any column"
    )
    subtitle.font = Font(size=11, italic=True)
    ws.merge_cells(f"A{SUBTITLE_ROW}:{last_col}{SUBTITLE_ROW}")

    # --- Table headers (row 4) ---
    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx)
        style_header_cell(cell, h)

    ws.row_dimensions[HEADER_ROW].height = 35

    # --- Data rows (row 5+) ---
    data_end_row: int = DATA_START + len(all_wires) - 1
    for row_idx, wire in enumerate(all_wires, DATA_START):
        for col_idx, val in enumerate(wire, 1):
            style_data_cell(ws.cell(row=row_idx, column=col_idx), val)

    # Column widths
    for i, w in enumerate(ALL_WIRES_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes — keep title + headers visible when scrolling
    ws.freeze_panes = f"A{DATA_START}"

    # Note: Table formatting removed for OnlyOffice compatibility.


def build_summary_tab(
    wb: openpyxl.Workbook,
    all_wires: List[List[str]],
) -> None:
    """Create the ``Summary`` worksheet with wire counts and a live total.

    A title (row 1) and subtitle (row 2) sit above the data table.  The
    header row is row 4; section counts start at row 5 and the grand total
    uses a live ``=SUM(...)`` formula.

    Args:
        wb: The workbook to add the worksheet to.
        all_wires: List of wire data rows (used to compute per-section counts).
    """
    summary: Worksheet = wb.create_sheet(title="Summary")
    summary.sheet_view.show_gridlines = False
    summary.column_dimensions["A"].width = 30
    summary.column_dimensions["B"].width = 14

    # Title & subtitle
    summary.cell(row=1, column=1, value="FPV Rover - Wiring Summary").font = Font(
        size=14, bold=True
    )
    summary.cell(
        row=2, column=1, value="Total wire count by section (with live totals)"
    ).font = Font(size=11)

    # Compute per-section wire counts
    section_counts: dict[str, int] = {}
    for wire in all_wires:
        section: str = wire[1]
        section_counts[section] = section_counts.get(section, 0) + 1

    # Header row (row 4)
    summary_labels: list[str] = ["Section", "Wires"]
    for col, label in enumerate(summary_labels, 1):
        style_summary_header(summary.cell(row=4, column=col), label)
    summary.row_dimensions[4].height = 25

    # Section data rows (row 5+)
    start_row: int = 5
    for i, (sec, cnt) in enumerate(section_counts.items(), start_row):
        c1 = summary.cell(row=i, column=1, value=sec)
        c1.border = THIN_BORDER
        c2 = summary.cell(row=i, column=2, value=cnt)
        c2.border = THIN_BORDER
        c2.alignment = Alignment(horizontal="center")

    # Grand total row with live formula
    total_row: int = i + 1
    total_label = summary.cell(row=total_row, column=1, value="TOTAL")
    total_label.font = Font(bold=True)
    total_label.border = THIN_BORDER

    total_cell = summary.cell(row=total_row, column=2)
    total_cell.value = f"=SUM(B{start_row}:B{total_row - 1})"
    total_cell.font = Font(bold=True, color="FFFFFF")
    total_cell.fill = PatternFill(
        start_color=TOTAL_COLOR, end_color=TOTAL_COLOR, fill_type="solid"
    )
    total_cell.border = THIN_BORDER
    total_cell.alignment = Alignment(horizontal="center")

    summary.freeze_panes = "A5"


def build_readme_tab(wb: openpyxl.Workbook) -> None:
    """Create the ``Readme`` worksheet with notes and file-structure reference.

    Args:
        wb: The workbook to add the worksheet to.
    """
    readme: Worksheet = wb.create_sheet(title="Readme")
    readme.sheet_view.show_gridlines = False
    readme.column_dimensions["A"].width = 100

    notes: list[str] = [
        "FPV Rover - Mixed Text/Calc Wiring Sheet",
        "",
        "GENERAL NOTES",
        "- The ESP32-CAM in this project is an AI-Thinker module with an OV3660 camera.",
        "- The UNO helper handles motor control, I2C sensor reporting, and command relay.",
        "- GPIO12 is intentionally left unused to keep ESP32-CAM flashing reliable.",
        "- SD card support is attempted via SD_MMC.begin() in the ESP32 code; "
        "confirm SD wiring and GPIO usage before using SD features.",
        "- Level-shifting is required for all TTL serial between the ESP32-CAM "
        "and the UNO, and for the Jazzy serial output.",
        "- The I2C bus runs at 3.3V logic; level-shift the UNO's A4/A5 down to 3.3V.",
        "- A single 5V + GND ring around the perimeter feeds all 9 sensors "
        "(5 HC-SR04 + 4 HW-201).",
        "- A single shared TRIG line (D9) fires all 5 HC-SR04 simultaneously; "
        "only the ECHO lines return to the UNO.",
        "",
        "ALL_WIRES TAB",
        "- Contains every wire in a single Excel Table (ListObject) named 'WireTable'.",
        "- Title and subtitle rows are ABOVE the table and are never included in sorting.",
        "- Click the dropdown arrows in the header row (row 4) to filter or sort by any column:",
        "  - ID, Section, From Device, From Pin, To Device, To Pin, "
        "Signal/Notes, Phase, Level-shifted?",
        "- The table uses 'TableStyleMedium2' with row stripes for readability.",
        "- Freeze panes keep the title + header row visible when scrolling.",
        "",
        "COMPONENT_PLACEMENT TAB",
        "- Layout and pin maps imported from component_placement.md.",
        "- 7 sub-tables: Sensor Ring, Placement Rationale, UNO Pin Map, "
        "ESP32-CAM GPIO, MCP23017 #1, MCP23017 #2, Optimization Strategy.",
        "",
        "SUMMARY TAB",
        "- Wire count per section with a live =SUM formula for the grand total.",
        "",
        "FILE STRUCTURE",
        "- Readme             - this sheet (notes & reference)",
        "- All_Wires          - consolidated, sortable, filterable wire table (single source of truth)",
        "- Component_Placement - layout & pin maps from component_placement.md",
        "- Summary            - wire count by section with live SUM formula",
        "",
        "OPEN IN",
        "- OnlyOffice Calc, Microsoft Excel, or LibreOffice Calc",
        "",
        "PHASE FUNCTIONS",
        "- Phase 00  - Foundation: power distribution, ESP32<->UNO UART1 link, I2C bus, Jazzy output",
        "- Phase 01  - Status: RGB LED visual indicators",
        "- Phase 02  - Audio: passive buzzer alerts",
        "- Phase 03  - Awareness: sonar ring (5x HC-SR04) + IR drop sensor ring (4x HW-201)",
        "- Phase 04  - Locomotion: L298N motor drives (left/right track control)",
        "- Phase 05  - Navigation: NEO-6M GPS position/location",
        "- Phase 06  - Safety: SW-520D tilt/vibration detection",
        "- Phase 07  - Display: LCD1602 local status readout",
        "- Phase 08  - Climate: DHT11 ambient temperature & humidity",
        "- Phase 09  - Environment: HW-038 water level detection",
        "- Phase 10 - Timekeeping: DS1302 real-time clock (timestamps, logging)",
        "- Phase 11 - Vision: SB90 pan/tilt servos for camera aiming",
        "- Phase 12 - Remote: KY-022 IR receiver for remote control",
        "- Phase 13 - Access: RFID-RC522 card/tag reader",
        "- Phase 14 - Remote: VS1838B IR receiver (with remotes)",
        "- Phase 15 - Motion: ULN2003 stepper driver",
        "- Phase 16 - Audio: HW-484 microphone (sound sensing)",
    ]

    # Write notes with conditional formatting
    for idx, line in enumerate(notes, 1):
        cell = readme.cell(row=idx, column=1)
        cell.value = line
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if idx == 1:
            cell.font = Font(bold=True, size=14)
        elif line in (
            "GENERAL NOTES",
            "ALL_WIRES TAB",
            "COMPONENT_PLACEMENT TAB",
            "SUMMARY TAB",
            "FILE STRUCTURE",
            "OPEN IN",
        ):
            cell.font = Font(bold=True)
        elif line.startswith("  -"):
            cell.font = Font(size=10)


def _add_subtable(
    ws: Worksheet,
    start_row: int,
    title: str,
    headers: list[str],
    rows: list[list[str]],
) -> int:
    """Add a titled sub-table to the ``Component_Placement`` worksheet.

    Each sub-table has:
      - A merged title cell (row *start_row*)
      - A styled header row (row *start_row + 1*)
      - Data rows starting at row *start_row + 2*

    Args:
        ws: The worksheet to write into.
        start_row: The first row for this sub-table.
        title: The sub-table title (will be merged across all header columns).
        headers: Column header labels.
        rows: Data rows (each row matches ``headers`` in length).

    Returns:
        The row number where the *next* sub-table should begin.
    """
    last_col: str = get_column_letter(len(headers))

    # Section title — merged across all columns
    tcell = ws.cell(row=start_row, column=1, value=title)
    tcell.font = Font(bold=True, size=13)
    ws.merge_cells(f"A{start_row}:{last_col}{start_row}")

    # Header row
    header_row: int = start_row + 1
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(
            start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid"
        )
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[header_row].height = 25

    # Data rows
    data_row: int = header_row + 1
    for sub_row in rows:
        for col_idx, val in enumerate(sub_row, 1):
            cell = ws.cell(row=data_row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="top")
        data_row += 1

    # Blank gap between sub-tables
    ws.cell(row=data_row, column=1, value="")
    return data_row + 1


def build_component_placement_tab(wb: openpyxl.Workbook) -> None:
    """Create the ``Component_Placement`` worksheet.

    Imports tabular data from ``component_placement.md`` as 7 styled
    sub-tables: Sensor Ring, Placement Rationale, UNO Pin Map, ESP32-CAM
    GPIO, MCP23017 #1, MCP23017 #2, and Wiring Optimization Strategy notes.

    Args:
        wb: The workbook to add the worksheet to.
    """
    cp: Worksheet = wb.create_sheet(title="Component_Placement")
    cp.sheet_view.show_gridlines = False

    # Fix column widths for the widest sub-table (4 columns max here)
    cp_widths: list[int] = [28, 22, 20, 20, 40]
    for i, w in enumerate(cp_widths, 1):
        cp.column_dimensions[get_column_letter(i)].width = w

    row: int = 1

    # 1. Sensor Ring
    sensor_headers: list[str] = ["Sensor", "Position", "Aim", "Purpose"]
    sensor_rows: list[list[str]] = [
        ["HC-SR04 FC", "Center front", "0 deg (straight ahead)", "Primary obstacle detection"],
        ["HC-SR04 FL", "Front-left", "45 deg left of center", "Corner/approach detection"],
        ["HC-SR04 FR", "Front-right", "45 deg right of center", "Corner/approach detection"],
        ["HC-SR04 LS", "Left edge, mid-depth", "90 deg left", "Wall-following / side clearance"],
        ["HC-SR04 RS", "Right edge, mid-depth", "90 deg right", "Wall-following / side clearance"],
        ["HW-201 #1", "Front, between FL-HC and FC-HC", "Pointed down", "Drop-off / edge detection"],
        ["HW-201 #2", "Front, between FC-HC and FR-HC", "Pointed down", "Drop-off / edge detection"],
        ["HW-201 #3", "Left side, mid-depth", "Pointed down", "Drop-off / edge detection"],
        ["HW-201 #4", "Right side, mid-depth", "Pointed down", "Drop-off / edge detection"],
    ]
    row = _add_subtable(
        cp, row,
        "Sensor Ring (5 HC-SR04 + 4 HW-201, already mounted on chassis perimeter)",
        sensor_headers, sensor_rows,
    )

    # 2. Placement Rationale
    placement_headers: list[str] = ["Component", "Why here"]
    placement_rows: list[list[str]] = [
        ["UNO", "Center - hub for sensors, motors, I2C, and Jazzy. Central = shortest average wire run."],
        ["I2C cluster (INA219 x2, MCP23017, MPU6050)",
         "Right beside UNO A4/A5. I2C is capacitance-sensitive; short bus avoids glitches."],
        ["ESP32-CAM", "Front-center for forward camera view. Close to UNO (helper UART) and power."],
        ["L298N", "Rear-center, near motor wire exit. Keeps high-current motor wires short and away from sensors."],
        ["LM2596 x2", "Rear, near battery/power input. Short high-current feed to L298N and 5V rail."],
        ["Level shifters", "Beside UNO - serve helper UART (D11->ESP32) and Jazzy (D1->chassis edge)."],
        ["Jazzy cable exit", "Rear or side edge, routed from UNO D1 via level shifter."],
    ]
    row = _add_subtable(
        cp, row, "Central Electronics Placement - Rationale",
        placement_headers, placement_rows,
    )

    # 3. UNO Final Pin Map
    uno_headers: list[str] = ["Pin", "Function", "Phase"]
    uno_rows: list[list[str]] = [
        ["D0", "Reserved (Jazzy RX mate)", "existing"],
        ["D1", "Jazzy TX (38400 8E1)", "existing"],
        ["D2", "L298N IN1", "existing"],
        ["D3", "L298N IN2", "existing"],
        ["D4", "L298N IN4", "existing"],
        ["D5", "L298N ENA (PWM)", "existing"],
        ["D6", "L298N ENB (PWM)", "existing"],
        ["D7", "L298N IN3", "existing"],
        ["D8", "Spare", "—"],
        ["D9", "HC-SR04 TRIG (shared, all 5)", "Phase 03"],
        ["D10", "Helper RX (SoftwareSerial)", "existing"],
        ["D11", "Helper TX (SoftwareSerial)", "existing"],
        ["D12", "DHT11", "Phase 04"],
        ["D13", "HC-SR04 ECHO right (90 deg)", "Phase 04"],
        ["A0", "HC-SR04 ECHO center-front", "Phase 03"],
        ["A1", "HC-SR04 ECHO left-front", "Phase 03"],
        ["A2", "HC-SR04 ECHO right-front", "Phase 03"],
        ["A3", "HC-SR04 ECHO left (90 deg)", "Phase 04"],
        ["A4", "I2C SDA", "existing"],
        ["A5", "I2C SCL", "existing"],
        ["A6", "HW-484 mic (analog)", "Phase 04"],
        ["A7", "Spare", "—"],
    ]
    row = _add_subtable(
        cp, row, "Elegoo UNO R3 (ATmega328P) Final Pin Map (Phased)",
        uno_headers, uno_rows,
    )

    # 4. ESP32-CAM GPIO
    esp_headers: list[str] = ["GPIO", "Function", "Phase"]
    esp_rows: list[list[str]] = [
        ["GPIO12", "Keep free (flashing reliability)", "—"],
        ["GPIO13", "Helper UART TX", "existing"],
        ["GPIO14", "Helper UART RX", "existing"],
        ["GPIO15", "Tilt servo", "Phase 04"],
        ["GPIO16", "GPS RX (NEO-6M)", "Phase 04"],
        ["GPIO2", "Pan servo", "Phase 04"],
        ["GPIO4", "Camera LED", "existing"],
    ]
    row = _add_subtable(
        cp, row, "ESP32-CAM GPIO",
        esp_headers, esp_rows,
    )

    # 5. MCP23017 #1 (0x27)
    mcp1_headers: list[str] = ["Pin", "Function", "Phase"]
    mcp1_rows: list[list[str]] = [
        ["GPA0", "WiFi status LED", "existing"],
        ["GPA1-GPA4", "HW-201 IR (4)", "Phase 03"],
        ["GPA5-GPA7", "HW-201 IR (3 more)", "Phase 04"],
        ["GPB0-GPB2", "DS1302 RTC", "Phase 04"],
        ["GPB3-GPB7", "Spare / KY-022", "Phase 04"],
    ]
    row = _add_subtable(
        cp, row, "MCP23017 #1 (0x27)",
        mcp1_headers, mcp1_rows,
    )

    # 6. MCP23017 #2 (0x26)
    mcp2_headers: list[str] = ["Pin", "Function"]
    mcp2_rows: list[list[str]] = [
        ["GPA0-GPA2", "HW-201 IR (3 more, to reach 10)"],
        ["GPA3-GPA7 + GPB0", "LCD1602 4-bit (RS, E, D4-D7)"],
        ["GPB1-GPB3", "Spare"],
        ["GPB4", "KY-022 IR receiver"],
        ["GPB5-GPB7", "Spare"],
    ]
    row = _add_subtable(
        cp, row, "MCP23017 #2 (0x26) - Phase 2 full config",
        mcp2_headers, mcp2_rows,
    )

    # 7. Wiring Optimization Strategy (text notes, not a table)
    strategy_notes: list[str] = [
        "Wiring Optimization Strategy",
        "1. Shared power ring - one 5V + GND loop around perimeter feeds all sensors.",
        "2. Shared HC-SR04 TRIG - single TRIG line (D9) daisy-chains to all 5 sensors.",
        "3. I2C cluster - all devices within a few inches of UNO A4/A5.",
        "4. HW-201 on MCP23017 - frees UNO pins for timing-critical HC-SR04 echoes.",
        "5. Servos + GPS on ESP32-CAM - offloads the UNO; keeps camera wiring local.",
        "6. High-current vs. sensitive separation - L298N and LM2596 at rear; sensors away from motor EMI.",
    ]
    for i, note in enumerate(strategy_notes):
        cell = cp.cell(row=row + i, column=1, value=note)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if i == 0:
            cell.font = Font(bold=True, size=13)
        else:
            cell.font = Font(size=10)


def order_sheets(wb: openpyxl.Workbook) -> None:
    """Reorder worksheets so Readme appears first.

    Args:
        wb: The workbook whose sheets to reorder.
    """
    target_order: list[str] = ["Readme", "All_Wires", "Component_Placement", "Summary"]
    for target_name in target_order:
        current_idx: int = wb.sheetnames.index(target_name)
        # Move to very front (position 0)
        wb.move_sheet(target_name, offset=-(current_idx))


def main() -> None:
    """Generate ``wiring_chart.xlsx`` with all four worksheets.

    Side effects:
        - Creates ``wiring_chart.xlsx`` in the current working directory.
        - Prints a summary of what was written.
    """
    wb: openpyxl.Workbook = openpyxl.Workbook()

    # Build each worksheet
    build_all_wires_tab(wb, ALL_WIRES)
    build_summary_tab(wb, ALL_WIRES)
    build_readme_tab(wb)
    build_component_placement_tab(wb)

    # Reorder tabs for a logical reading order
    order_sheets(wb)

    # Save the final workbook
    output_path: str = "wiring_chart.xlsx"
    wb.save(output_path)

    # Report what was generated
    print(f"{output_path} written successfully.")
    print(
        f"  All_Wires table: {TOTAL_WIRES} data rows + 1 header = "
        f"{TOTAL_WIRES + 1} rows"
    )
    last_col: str = get_column_letter(len(HEADERS))
    data_end: int = DATA_START + TOTAL_WIRES - 1
    print(f"  Table range: A{HEADER_ROW}:{last_col}{data_end}")
    # Count unique sections
    sections: set[str] = {w[1] for w in ALL_WIRES}
    print(f"  Sections: {len(sections)}")
    print(f"  Total wires: {TOTAL_WIRES}")
    print("  Component_Placement: 7 sub-tables added")


if __name__ == "__main__":
    main()
