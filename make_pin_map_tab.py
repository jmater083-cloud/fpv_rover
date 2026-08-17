#!/usr/bin/env python3
"""
Modify wiring_chart.xlsx:
  - Remove the Summary tab.
  - Create a new consolidated "Pin_Map" tab built from the pin/GPIO data
    in the Component_Placement tab.

The new Pin_Map tab has four columns:
    Section | Pin/GPIO | Function | Phase

Sections included (from Component_Placement):
    - UNO (ATmega328P) Final Pin Map
    - ESP32-CAM GPIO
    - MCP23017 #1 (0x27)
    - MCP23017 #2 (0x26)  (Phase filled as "Phase 2" per its section title)
"""

from __future__ import annotations

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

INPUT_PATH: str = "wiring_chart.xlsx"

# ── Styling constants (match make_wiring_xlsx.py) ────────────────────────────
THIN_BORDER: Border = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
HEADER_COLOR: str = "1F4E78"

#: Column headers for the consolidated Pin_Map table.
HEADERS: list[str] = ["Section", "Pin/GPIO", "Function", "Phase"]

#: Column widths (characters) matching HEADERS order.
WIDTHS: list[int] = [16, 14, 45, 12]

#: Section display names keyed by the Component_Placement section title.
SECTION_NAMES: dict[str, str] = {
    "Elegoo UNO R3 (ATmega328P) Final Pin Map (Phased)": "UNO",
    "ESP32-CAM GPIO": "ESP32-CAM",
    "MCP23017 #1 (0x27)": "MCP23017 #1",
    "MCP23017 #2 (0x26) - Phase 2 full config": "MCP23017 #2",
}

#: Sections whose rows have a Phase column (Pin, Function, Phase).
PHASE_SECTIONS: set[str] = {
    "Elegoo UNO R3 (ATmega328P) Final Pin Map (Phased)",
    "ESP32-CAM GPIO",
    "MCP23017 #1 (0x27)",
}

#: Sections without a Phase column; Phase is filled from this value.
DEFAULT_PHASE: str = "Phase 2"

#: Rows to expand into individual pin rows.
#: Keyed by (section, pin, function) -> list of (pin, function) replacements.
#: Used to break a grouped pin range into per-pin rows with specific positions.
EXPANSIONS: dict[tuple[str, str, str], list[tuple[str, str]]] = {
    (
        "MCP23017 #1",
        "GPA1-GPA4",
        "HW-201 IR (4)",
    ): [
        ("GPA1", "HW-201 IR (front-left)"),
        ("GPA2", "HW-201 IR (front-right)"),
        ("GPA3", "HW-201 IR (left)"),
        ("GPA4", "HW-201 IR (right)"),
    ],
}


def collect_pin_rows(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[list[str]]:
    """Extract consolidated pin rows from the Component_Placement worksheet.

    Scans for section-title rows (merged, single value in column A) that match
    SECTION_NAMES, then reads the data rows that follow until the next blank
    row or section title.

    Args:
        ws: The Component_Placement worksheet.

    Returns:
        A list of rows, each ``[Section, Pin, Function, Phase]``.
    """
    rows: list[list[str]] = []
    current_section: str | None = None
    has_phase: bool = False

    for row in ws.iter_rows():
        # First cell value drives section detection
        first_val = row[0].value
        if first_val in SECTION_NAMES:
            current_section = SECTION_NAMES[first_val]
            has_phase = first_val in PHASE_SECTIONS
            continue

        if current_section is None:
            continue

        # Stop at blank row or a new section title
        if first_val is None or first_val in SECTION_NAMES:
            current_section = None
            continue

        # Skip header rows (e.g. "Pin", "GPIO")
        if first_val in ("Pin", "GPIO"):
            continue

        pin = str(first_val).strip()
        function = str(row[1].value).strip() if row[1].value else ""
        if has_phase:
            phase = str(row[2].value).strip() if row[2].value else "—"
        else:
            phase = DEFAULT_PHASE

        # Expand grouped pin ranges into individual rows (e.g. GPA1-GPA4)
        expansion = EXPANSIONS.get((current_section, pin, function))
        if expansion:
            for exp_pin, exp_function in expansion:
                rows.append([current_section, exp_pin, exp_function, phase])
        else:
            rows.append([current_section, pin, function, phase])

    return rows


def build_pin_map_tab(
    wb: openpyxl.Workbook,
    pin_rows: list[list[str]],
) -> None:
    """Create the consolidated ``Pin_Map`` worksheet.

    Args:
        wb: The workbook to add the worksheet to.
        pin_rows: Consolidated rows ``[Section, Pin, Function, Phase]``.
    """
    ws = wb.create_sheet(title="Pin_Map")
    ws.sheet_view.show_gridlines = False

    last_col: str = get_column_letter(len(HEADERS))

    # Title (row 1)
    title = ws.cell(row=1, column=1, value="FPV Rover - Pin / GPIO Map")
    title.font = Font(size=14, bold=True)
    ws.merge_cells(f"A1:{last_col}1")

    # Subtitle (row 2)
    subtitle = ws.cell(
        row=2,
        column=1,
        value="Consolidated pin assignments from Component_Placement",
    )
    subtitle.font = Font(size=11, italic=True)
    ws.merge_cells(f"A2:{last_col}2")

    # Header row (row 4)
    header_row: int = 4
    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(
            start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[header_row].height = 25

    # Data rows (row 5+)
    data_start: int = header_row + 1
    for r_idx, pin_row in enumerate(pin_rows, data_start):
        for c_idx, val in enumerate(pin_row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="top")

    # Column widths
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes — keep title + headers visible
    ws.freeze_panes = f"A{data_start}"


def main() -> None:
    """Load the workbook, remove Summary, add Pin_Map, and save."""
    wb: openpyxl.Workbook = openpyxl.load_workbook(INPUT_PATH)

    # Remove the Summary tab if present
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
        print("Removed 'Summary' tab.")

    # Remove an existing Pin_Map tab so it can be rebuilt
    if "Pin_Map" in wb.sheetnames:
        del wb["Pin_Map"]
        print("Removed existing 'Pin_Map' tab.")

    # Build Pin_Map from Component_Placement data
    cp: openpyxl.worksheet.worksheet.Worksheet = wb["Component_Placement"]
    pin_rows: list[list[str]] = collect_pin_rows(cp)
    build_pin_map_tab(wb, pin_rows)
    print(f"Created 'Pin_Map' tab with {len(pin_rows)} rows.")

    # Move Pin_Map to a sensible position (after Component_Placement)
    wb.move_sheet("Pin_Map", offset=len(wb.sheetnames) - 1 - wb.sheetnames.index("Pin_Map"))

    wb.save(INPUT_PATH)
    print(f"Saved {INPUT_PATH}. Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()